import openpyxl

#------Method to read data from the excel sheet-------#
def readExcelData():

    bearingsList=[]
    workbook=openpyxl.load_workbook("../testData/Bearings.xlsx")
    sheet=workbook.active

    for i in range (1,sheet.max_row+1):
        for j in range (1,sheet.max_column+1):
            bearingType = (sheet.cell(row=i,column=j).value).rstrip()
            bearingsList.append(bearingType)

    return bearingsList

